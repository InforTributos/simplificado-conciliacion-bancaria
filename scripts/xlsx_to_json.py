"""Convierte los archivos movimientos.xlsx de cada carpeta en tests/fixtures/reales-completas/
a movimientos.json con la estructura esperada por la API /api/v1/conciliaciones/procesar.

Uso:
    python scripts/xlsx_to_json.py
"""

import json
import os
from datetime import datetime

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIXTURES_DIR = os.path.join(BASE_DIR, "tests", "fixtures", "reales-completas")

FOLDERS = ["Cartagena", "Monteria", "Valledupar"]


def to_aaaamm_format(dt):
    """Convert datetime to dd-mm-aaaa string."""
    if isinstance(dt, datetime):
        return dt.strftime("%d-%m-%Y")
    if isinstance(dt, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(dt.strip(), fmt).strftime("%d-%m-%Y")
            except ValueError:
                continue
    return str(dt)


def is_data_row(row_values):
    """Check if row contains actual movement data (has code, type, etc)."""
    if not row_values:
        return False
    first_cell = str(row_values[0]).strip().upper() if row_values[0] is not None else ""
    if first_cell in ("IRM", "NCO", "CE", "EGR", "NCB"):
        return True
    return False


def extract_movements(ws):
    """Extract movements from worksheet, auto-detecting column positions."""
    headers = {}
    header_row = -1

    # Find header row (row with TIPO / Tipo in col A)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        first_cell = str(row[0]).strip().upper() if row[0] is not None else ""
        if first_cell in ("TIPO", "CODIGO"):
            # Try row as header; if not, next row
            if first_cell == "TIPO":
                header_row = i
                for j, h in enumerate(row):
                    if h is not None:
                        headers[str(h).strip().upper()] = j
                break
            elif first_cell == "CODIGO":
                # This might be a "Codigo:" row or actual data
                if str(row[1]).strip() == "1908030101":
                    # It's the "Codigo:" account row, skip
                    continue
                # Otherwise try the previous row as header (Monteria style)
                for j, h in enumerate(row):
                    if h is not None:
                        headers[str(h).strip().upper()] = j
                header_row = i
                break

    if not headers:
        return []

    # Now detect which columns to use based on header names
    col_tipo = headers.get("TIPO", 0)
    col_codigo = headers.get("CODIGO", headers.get("CODIGO", 1))
    col_fecha = headers.get("FECHA", 2)
    col_debito = headers.get("DEBITO", 7)
    col_credito = headers.get("CREDITO", 8)
    col_saldo = None
    for key in ("SALDO FINAL", "SALDO"):
        if key in headers:
            col_saldo = headers[key]
            break
    if col_saldo is None:
        col_saldo = 9

    movements = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)

        # Detect end-of-data markers (check any cell in the row)
        row_text = " ".join(str(v).strip().upper() for v in vals if v is not None)
        if any(
            m in row_text
            for m in ("SUMAS IGUALES", "SALDO FINAL", "TOTAL REGISTROS",
                       "GENERADO POR", "FECHA Y HORA")
        ):
            break
        if not is_data_row(vals):
            continue

        # Pad row if shorter than needed
        max_col = max(col_tipo, col_codigo, col_fecha, col_debito, col_credito, col_saldo)
        while len(vals) <= max_col:
            vals.append(None)

        tipo = str(vals[col_tipo]).strip() if vals[col_tipo] is not None else ""
        codigo = str(vals[col_codigo]).strip() if vals[col_codigo] is not None else ""
        codigo = codigo.replace(".0", "") if codigo.endswith(".0") else codigo

        fecha_val = vals[col_fecha]
        fecha_str = to_aaaamm_format(fecha_val)

        def safe_float(v, default=0.0):
            if v is None:
                return default
            try:
                return float(v)
            except (ValueError, TypeError):
                return default

        debito = safe_float(vals[col_debito])
        credito = safe_float(vals[col_credito])
        saldo = safe_float(vals[col_saldo])

        movements.append({
            "fecha": fecha_str,
            "codigo_movimiento": codigo,
            "debito": debito,
            "credito": credito,
            "saldo": saldo,
            "conciliado": False,
        })

    return movements


def cols_trim(vals, col):
    """Get trimmed string from column index safely."""
    if col < len(vals) and vals[col] is not None:
        return str(vals[col]).strip()
    return ""


def process_folder(folder_name):
    xlsx_path = os.path.join(FIXTURES_DIR, folder_name, "movimientos.xlsx")
    json_path = os.path.join(FIXTURES_DIR, folder_name, "movimientos.json")

    if not os.path.exists(xlsx_path):
        print(f"[SKIP] {xlsx_path} no existe")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    movements = extract_movements(ws)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(movements, f, indent=2, ensure_ascii=False)

    print(f"[OK] {folder_name}: {len(movements)} movimientos -> {json_path}")
    return movements


def main():
    for folder in FOLDERS:
        process_folder(folder)


if __name__ == "__main__":
    main()
